import cv2
import numpy as np
from ultralytics import YOLO
import math
from datetime import datetime, timedelta
import json
import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# YOLO modelini yuklash
model = YOLO("yolov8m.pt")

# Berilgan to'rtburchak koordinatalari va ismlar
rectangles = [
    ("Bahrombek", (861, 784, 332, 427)),
    ("Islom", (1284, 777, 366, 419)),
    ("Sardor", (887, 452, 291, 208)),
    ("Javohir", (1201, 493, 230, 196))
]

def is_person_in_area(person_box, area_box):
    px1, py1, px2, py2 = person_box
    ax, ay, aw, ah = area_box
    ax1, ay1, ax2, ay2 = ax, ay, ax + aw, ay + ah
    return (ax1 < px1 < ax2 or ax1 < px2 < ax2) and (ay1 < py1 < ay2 or ay1 < py2 < ay2)

def save_time_data(total_times, start_times):
    data = {
        name: {
            "total_time": total_time.total_seconds(),
            "start_time": start_time.isoformat() if start_time else None
        } for (name, _), (total_time, start_time) in zip(rectangles, zip(total_times, start_times))
    }
    with open("time_data.json", "w") as f:
        json.dump(data, f)

def load_time_data():
    if os.path.exists("time_data.json"):
        try:
            with open("time_data.json", "r") as f:
                data = json.load(f)
            total_times = []
            start_times = []
            for name, _ in rectangles:
                area_data = data.get(name, {"total_time": 0, "start_time": None})
                total_times.append(timedelta(seconds=area_data["total_time"]))
                start_times.append(datetime.fromisoformat(area_data["start_time"]) if area_data["start_time"] else None)
            return total_times, start_times
        except (json.JSONDecodeError, KeyError, ValueError):
            print("Saqlangan ma'lumotlarni o'qishda xatolik. Yangi ma'lumotlar yaratilmoqda.")
    return [timedelta() for _ in range(len(rectangles))], [None for _ in range(len(rectangles))]

def update_excel(total_times):
    file_name = 'time_tracking.xlsx'
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Tracking"
        ws.append(["Vaqt"] + [name for name, _ in rectangles])

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [current_time]
    for total_time in total_times:
        hours, remainder = divmod(int(total_time.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)
        row.append(f"{hours:02d}:{minutes:02d}:{seconds:02d}")
    
    ws.append(row)

    # Ustun kengligini moslash
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_name)

def main():
    cap = cv2.VideoCapture("rtsp://admin:DAS2024@@192.168.136.234:554/Streamin/Channels/401")
    
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter('output3.mp4', fourcc, 20.0, (width, height))
    
    total_times, start_times = load_time_data()
    persons_in_areas = [start_time is not None for start_time in start_times]

    last_excel_update = datetime.now()

    try:
        while True:
            success, frame = cap.read()
            if not success:
                print("Kamera bilan bog'lanishda xatolik. Qayta urinish...")
                cap.release()
                time.sleep(5)
                cap = cv2.VideoCapture("rtsp://admin:DAS2024@@192.168.136.234:554/Streamin/Channels/401")
                continue

            results = model(frame, stream=True)

            persons_detected = [False for _ in range(len(rectangles))]
            for r in results:
                boxes = r.boxes
                for box in boxes:
                    if int(box.cls[0]) == 0:  # 0 - odam sinfi
                        x_center, y_center, width, height = box.xywh[0]
                        x1 = int(x_center - width/2)
                        y1 = int(y_center - height/2)
                        x2 = int(x_center + width/2)
                        y2 = int(y_center + height/2)

                        for i, (_, rect) in enumerate(rectangles):
                            if is_person_in_area((x1, y1, x2, y2), rect):
                                persons_detected[i] = True
                                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                                conf = math.ceil(box.conf[0] * 100) / 100
                                cv2.putText(frame, f"Human: {conf}", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

            for i, ((name, (x, y, w, h)), person_detected) in enumerate(zip(rectangles, persons_detected)):
                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
                cv2.putText(frame, name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)

            current_time = datetime.now()
            for i in range(len(rectangles)):
                if persons_detected[i] and not persons_in_areas[i]:
                    start_times[i] = current_time
                    persons_in_areas[i] = True
                elif not persons_detected[i] and persons_in_areas[i]:
                    total_times[i] += current_time - start_times[i]
                    start_times[i] = None
                    persons_in_areas[i] = False

                if persons_in_areas[i]:
                    current_duration = current_time - start_times[i] + total_times[i]
                else:
                    current_duration = total_times[i]

                hours, remainder = divmod(int(current_duration.total_seconds()), 3600)
                minutes, seconds = divmod(remainder, 60)
                time_str = f"{rectangles[i][0]}: {hours:02d}:{minutes:02d}:{seconds:02d}"
                cv2.putText(frame, time_str, (20, 40 + i*40), cv2.FONT_HERSHEY_SIMPLEX, 1.3, (0, 0, 255), 3)

            out.write(frame)

            cv2.namedWindow("Human Detection", cv2.WINDOW_NORMAL)
            cv2.resizeWindow("Human Detection", 640, 480)
            cv2.imshow("Human Detection", frame)

            save_time_data(total_times, start_times)

            # Har minutda Excel faylini yangilash
            if (current_time - last_excel_update).total_seconds() >= 60:
                update_excel(total_times)
                last_excel_update = current_time

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    finally:
        cap.release()
        out.release()
        cv2.destroyAllWindows()

if __name__ == "__main__":
    while True:
        try:
            main()
        except Exception as e:
            print(f"Xatolik yuz berdi: {e}")
            print("Dastur qayta ishga tushirilmoqda...")
            time.sleep(5)